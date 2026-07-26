# -*- coding: utf-8 -*-
"""
منظومة تنسيق رياض الأطفال والمدارس الرسمية للغات
مديرية التربية والتعليم بأسوان - محافظة أسوان
--------------------------------------------------
تطبيق متكامل يدعم:
1. معالجة شيت الإكسيل العام وإنشاء شيتات مستقلة لكل مدرسة.
2. استخراج تقارير PDF المنسقة للطباعة لكل مدرسة.
3. واجهة مستخدم Kivy تدعم اللغة العربية بدون مربعات رموز.
"""

import os
import sys
import datetime
import calendar
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side

# ==========================================
# 1. استيراد مكتبات PDF واللغة العربية
# ==========================================
try:
    from fpdf import FPDF
    import arabic_reshaper
    from bidi.algorithm import get_display
    HAS_PDF_LIBS = True
except ImportError:
    HAS_PDF_LIBS = False

# ==========================================
# 2. استيراد Kivy وتسجيل الخط العربي
# ==========================================
try:
    from kivy.app import App
    from kivy.uix.boxlayout import BoxLayout
    from kivy.uix.gridlayout import GridLayout
    from kivy.uix.label import Label
    from kivy.uix.textinput import TextInput
    from kivy.uix.button import Button
    from kivy.core.text import LabelBase
    HAS_KIVY = True

    # تسجيل الخط العربي لحل مشكلة المربعات (☒☒☒) في Kivy
    ARABIC_FONT = "arial.ttf"
    if os.path.exists(ARABIC_FONT):
        LabelBase.register(name='Roboto', fn_regular=ARABIC_FONT)
    elif os.path.exists("C:/Windows/Fonts/arial.ttf"):
        LabelBase.register(name='Roboto', fn_regular="C:/Windows/Fonts/arial.ttf")
    elif os.path.exists("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"):
        LabelBase.register(name='Roboto', fn_regular="/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
except ImportError:
    HAS_KIVY = False


# ==========================================
# 3. دالة معالجة النص العربي للـ PDF والواجهة
# ==========================================
def fix_ar(text):
    if text is None:
        return ""
    text_str = str(text).strip()
    if not text_str:
        return ""
    if HAS_PDF_LIBS:
        reshaped = arabic_reshaper.reshape(text_str)
        return get_display(reshaped)
    return text_str


# ==========================================
# 4. دالة حساب السن في أول أكتوبر
# ==========================================
def calculate_age(dob_date, target_date=None):
    if target_date is None:
        target_date = datetime.date(2026, 10, 1)
    
    if isinstance(dob_date, datetime.datetime):
        dob_date = dob_date.date()
    elif isinstance(dob_date, str):
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y'):
            try:
                dob_date = datetime.datetime.strptime(dob_date.strip(), fmt).date()
                break
            except ValueError:
                pass
    
    if not isinstance(dob_date, datetime.date):
        return 0, 0, 0

    years = target_date.year - dob_date.year
    months = target_date.month - dob_date.month
    days = target_date.day - dob_date.day

    if days < 0:
        months -= 1
        prev_month = target_date.month - 1 if target_date.month > 1 else 12
        prev_year = target_date.year if target_date.month > 1 else target_date.year - 1
        _, days_in_prev = calendar.monthrange(prev_year, prev_month)
        days += days_in_prev

    if months < 0:
        years -= 1
        months += 12

    return years, months, days


# ==========================================
# 5. كلاس إنشاء تقرير ה-PDF للمدرسة
# ==========================================
class SchoolReportPDF(FPDF if HAS_PDF_LIBS else object):
    def __init__(self, school_name, stage_name, school_year, logo_path=None):
        if not HAS_PDF_LIBS:
            return
        super().__init__(orientation='P', unit='mm', format='A4')
        self.school_name = school_name
        self.stage_name = stage_name
        self.school_year = school_year
        self.logo_path = logo_path
        
        # اختيار الخط العربي
        font_path = "arial.ttf"
        font_bd_path = "arialbd.ttf"
        if not os.path.exists(font_path):
            if os.path.exists("C:/Windows/Fonts/arial.ttf"):
                font_path = "C:/Windows/Fonts/arial.ttf"
                font_bd_path = "C:/Windows/Fonts/arialbd.ttf"
            elif os.path.exists("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"):
                font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
                font_bd_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"

        try:
            self.add_font('ArabicFont', '', font_path, uni=True)
            self.add_font('ArabicFont', 'B', font_bd_path if os.path.exists(font_bd_path) else font_path, uni=True)
        except Exception as e:
            print(f"Font load warning: {e}")

    def header(self):
        if not HAS_PDF_LIBS:
            return
        self.set_font('ArabicFont', 'B', 11)
        
        # 1. الجزء الأيمن (المديرية والإدارة)
        self.set_xy(110, 10)
        self.cell(90, 6, fix_ar("مديرية التربية والتعليم بأسوان"), ln=1, align='R')
        self.set_x(110)
        self.cell(90, 6, fix_ar("إدارة المدارس الرسمية والمتميزة للغات"), ln=1, align='R')
        
        # 2. الجزء الأيسر (اللوجو)
        if self.logo_path and os.path.exists(self.logo_path):
            try:
                self.image(self.logo_path, x=12, y=10, w=25)
            except Exception as e:
                print(f"Logo error: {e}")
            
        self.ln(6)
        
        # 3. المنتصف (العنوان واسم المدرسة)
        self.set_font('ArabicFont', 'B', 12)
        title = f"كشف تنسيق المرحلة ({self.stage_name}) للعام الدراسي {self.school_year}"
        self.cell(0, 7, fix_ar(title), ln=1, align='C')
        
        school_str = f"اسم المدرسة : {self.school_name}"
        self.cell(0, 7, fix_ar(school_str), ln=1, align='C')
        self.ln(3)

    def draw_table_header(self):
        if not HAS_PDF_LIBS:
            return
        self.set_font('ArabicFont', 'B', 10)
        curr_y = self.get_y()
        
        # رسم الترويسة المدمجة (من اليمين لليسار)
        self.set_xy(10, curr_y)
        self.cell(44, 12, fix_ar("ملاحظات"), border=1, align='C')
        self.cell(36, 6, fix_ar("السن في أول أكتوبر"), border=1, align='C')
        self.cell(25, 12, fix_ar("تاريخ الميلاد"), border=1, align='C')
        self.cell(55, 12, fix_ar("اسم الطالب"), border=1, align='C')
        self.cell(10, 12, fix_ar("م"), border=1, align='C')
        
        # الصف الثاني لتقسيم السن
        self.set_xy(54, curr_y + 6)
        self.cell(12, 6, fix_ar("سنة"), border=1, align='C')
        self.cell(12, 6, fix_ar("شهر"), border=1, align='C')
        self.cell(12, 6, fix_ar("يوم"), border=1, align='C')
        
        self.set_xy(10, curr_y + 12)

    def footer_signatures(self):
        if not HAS_PDF_LIBS:
            return
        self.set_y(-28)
        self.set_font('ArabicFont', 'B', 11)
        
        col_w = 63
        self.set_x(10)
        # التوقيعات من اليمين لليسار
        self.cell(col_w, 6, fix_ar("يعتمد مدير المديرية"), border=0, align='C')
        self.cell(col_w, 6, fix_ar("مدير التعليم العام"), border=0, align='C')
        self.cell(col_w, 6, fix_ar("لجنة التنسيق"), border=0, align='C')


# ==========================================
# 6. إضافة شيتات المدارس لملف الإكسيل
# ==========================================
def add_school_sheets_to_excel(excel_path, output_excel_path, schools_data):
    wb = openpyxl.load_workbook(excel_path)
    
    header_font = Font(name='Arial', size=11, bold=True)
    data_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
    )

    for school_name, students in schools_data.items():
        clean_title = str(school_name)[:30].replace(":", "").replace("/", "-").replace("\\", "").replace("?", "").replace("*", "")
        
        if clean_title in wb.sheetnames:
            ws = wb[clean_title]
            ws.delete_rows(1, ws.max_row + 1)
        else:
            ws = wb.create_sheet(title=clean_title)
            
        ws.sheet_view.rightToLeft = True

        # إعداد هيدر الشيت
        ws.merge_cells('A1:A2')
        ws['A1'] = "م"
        
        ws.merge_cells('B1:B2')
        ws['B1'] = "اسم الطالب"
        
        ws.merge_cells('C1:C2')
        ws['C1'] = "تاريخ الميلاد"
        
        ws.merge_cells('D1:F1')
        ws['D1'] = "السن في أول أكتوبر"
        
        ws['D2'] = "سنة"
        ws['E2'] = "شهر"
        ws['F2'] = "يوم"
        
        ws.merge_cells('G1:G2')
        ws['G1'] = "ملاحظات"

        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=7):
            for cell in row:
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

        col_widths = {'A': 6, 'B': 32, 'C': 15, 'D': 8, 'E': 8, 'F': 8, 'G': 25}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # إضافة الطلاب
        for idx, student in enumerate(students, 1):
            r = idx + 2
            ws.cell(row=r, column=1, value=idx).alignment = center_align
            ws.cell(row=r, column=2, value=student.get('name', '')).alignment = right_align
            ws.cell(row=r, column=3, value=student.get('dob', '')).alignment = center_align
            ws.cell(row=r, column=4, value=student.get('age_years', 0)).alignment = center_align
            ws.cell(row=r, column=5, value=student.get('age_months', 0)).alignment = center_align
            ws.cell(row=r, column=6, value=student.get('age_days', 0)).alignment = center_align
            ws.cell(row=r, column=7, value=student.get('notes', '')).alignment = right_align
            
            for col_idx in range(1, 8):
                c = ws.cell(row=r, column=col_idx)
                c.font = data_font
                c.border = thin_border

    wb.save(output_excel_path)


# ==========================================
# 7. توليد تقارير PDF المدارس
# ==========================================
def generate_all_pdf_reports(schools_data, stage_name, school_year, output_dir, logo_path=None):
    if not HAS_PDF_LIBS:
        return
        
    os.makedirs(output_dir, exist_ok=True)
    
    for school_name, students in schools_data.items():
        safe_school_name = str(school_name).replace(":", "").replace("/", "-").replace("\\", "")
        pdf_filename = os.path.join(output_dir, f"كشف_{safe_school_name}_المرحلة_{stage_name}.pdf")
        
        pdf = SchoolReportPDF(school_name, stage_name, school_year, logo_path)
        pdf.add_page()
        pdf.draw_table_header()
        
        pdf.set_font('ArabicFont', '', 10)
        
        for idx, student in enumerate(students, 1):
            if pdf.get_y() > 240:
                pdf.add_page()
                pdf.draw_table_header()
                pdf.set_font('ArabicFont', '', 10)
                
            curr_y = pdf.get_y()
            
            # كتابة السطر (يمين إلى يسار)
            pdf.set_xy(10, curr_y)
            pdf.cell(44, 7, fix_ar(student.get('notes', '')), border=1, align='R')
            pdf.cell(12, 7, str(student.get('age_years', '')), border=1, align='C')
            pdf.cell(12, 7, str(student.get('age_months', '')), border=1, align='C')
            pdf.cell(12, 7, str(student.get('age_days', '')), border=1, align='C')
            pdf.cell(25, 7, str(student.get('dob', '')), border=1, align='C')
            pdf.cell(55, 7, fix_ar(student.get('name', '')), border=1, align='R')
            pdf.cell(10, 7, str(idx), border=1, align='C', ln=1)

        pdf.footer_signatures()
        pdf.output(pdf_filename)


# ==========================================
# 8. المعالجة الرئيسية
# ==========================================
def process_coordination(excel_path, logo_path, stage_name="الأولى", school_year="2026/2027", target_date_str="2026-10-01"):
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"لم يتم العثور على الملف: {excel_path}")

    target_date = datetime.datetime.strptime(target_date_str, "%Y-%m-%d").date()
    
    wb = openpyxl.load_workbook(excel_path)
    sheet_name = "الطلاب" if "الطلاب" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    schools_data = {}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
            
        student_name = row[1] if len(row) > 1 else ""
        dob_raw = row[2] if len(row) > 2 else ""
        assigned_school = row[3] if len(row) > 3 else "قائمة الانتظار"
        notes = row[4] if len(row) > 4 else "مقبول تساوي سن"
        
        if not student_name:
            continue
            
        y, m, d = calculate_age(dob_raw, target_date)
        dob_str = dob_raw.strftime("%Y-%m-%d") if isinstance(dob_raw, (datetime.date, datetime.datetime)) else str(dob_raw)
        
        student_dict = {
            'name': student_name,
            'dob': dob_str,
            'age_years': y,
            'age_months': m,
            'age_days': d,
            'notes': notes
        }
        
        school_key = str(assigned_school).strip() if assigned_school else "قائمة الانتظار"
        if school_key not in schools_data:
            schools_data[school_key] = []
            
        schools_data[school_key].append(student_dict)

    dir_name = os.path.dirname(excel_path) or "."
    output_excel_name = f"منظومة_التنسيق_المرحلة_{stage_name}.xlsx"
    output_excel_path = os.path.join(dir_name, output_excel_name)
    
    # 1. تحديث الإكسيل وإضافة شيتات المدارس
    add_school_sheets_to_excel(excel_path, output_excel_path, schools_data)
    
    # 2. توليد كشوف ה-PDF
    output_pdf_dir = os.path.join(dir_name, f"كشوف_المدارس_المرحلة_{stage_name}_PDF")
    generate_all_pdf_reports(schools_data, stage_name, school_year, output_pdf_dir, logo_path)
    
    return output_excel_path, output_pdf_dir


# ==========================================
# 9. واجهة Kivy المحدثة لدعم العربية
# ==========================================
if HAS_KIVY:
    class KGCoordinationGUI(BoxLayout):
        def __init__(self, **kwargs):
            super().__init__(**kwargs)
            self.orientation = 'vertical'
            self.padding = 15
            self.spacing = 10
            
            # العناوين مع المعالجة العربية
            self.add_widget(Label(text=fix_ar("مديرية التربية والتعليم بأسوان"), font_size='20sp', size_hint_y=None, height=35))
            self.add_widget(Label(text=fix_ar("منظومة تنسيق رياض الأطفال والمدارس الرسمية للغات"), font_size='15sp', size_hint_y=None, height=30))
            
            # حقول الإدخال
            grid = GridLayout(cols=2, spacing=10, size_hint_y=None, height=180)
            
            grid.add_widget(Label(text=fix_ar("مسار ملف الإكسيل:")))
            self.excel_input = TextInput(text="students.xlsx", multiline=False)
            grid.add_widget(self.excel_input)
            
            grid.add_widget(Label(text=fix_ar("مسار صورة اللوجو (اختياري):")))
            self.logo_input = TextInput(text="logo.png", multiline=False)
            grid.add_widget(self.logo_input)
            
            grid.add_widget(Label(text=fix_ar("اسم المرحلة:")))
            self.stage_input = TextInput(text="الأولى", multiline=False)
            grid.add_widget(self.stage_input)
            
            grid.add_widget(Label(text=fix_ar("العام الدراسي:")))
            self.year_input = TextInput(text="2026/2027", multiline=False)
            grid.add_widget(self.year_input)
            
            self.add_widget(grid)
            
            # زر التشغيل
            self.btn_run = Button(
                text=fix_ar("بدء التنسيق واستخراج التقارير"), 
                background_color=(0.2, 0.6, 0.2, 1), 
                font_size='18sp', 
                size_hint_y=None, 
                height=50
            )
            self.btn_run.bind(on_press=self.run_process)
            self.add_widget(self.btn_run)
            
            # منطقة المخرجات والسجلات
            self.log_area = TextInput(readonly=True, text=fix_ar("جاهز لبدء العمل...\n"))
            self.add_widget(self.log_area)

        def log(self, msg):
            self.log_area.text += fix_ar(str(msg)) + "\n"

        def run_process(self, instance):
            try:
                excel_p = self.excel_input.text.strip()
                logo_p = self.logo_input.text.strip()
                stage_p = self.stage_input.text.strip()
                year_p = self.year_input.text.strip()
                
                self.log("جاري معالجة البيانات واستخراج التقارير...")
                out_excel, out_pdf = process_coordination(excel_p, logo_p if os.path.exists(logo_p) else None, stage_p, year_p)
                
                self.log(f"تم بنجاح! تم حفظ ملف الإكسيل:\n{out_excel}")
                self.log(f"تم حفظ تقارير الـ PDF في المجلد:\n{out_pdf}")
            except Exception as e:
                self.log(f"حدث خطأ أثناء المعالجة: {e}")

    class KGApp(App):
        def build(self):
            self.title = fix_ar("منظومة تنسيق رياض الأطفال - أسوان")
            return KGCoordinationGUI()

if __name__ == "__main__":
    if HAS_KIVY:
        KGApp().run()
    else:
        print("Kivy is not installed. Running backend logic...")
